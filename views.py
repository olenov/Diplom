from main import *
from models import *

from flask_mail import Mail
from flask import render_template, request, redirect, url_for, session, send_from_directory, flash, jsonify
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField
from urllib.parse import urlparse, urljoin
import os, xlsxwriter, string, random
from flask_login import login_user, login_required, logout_user, current_user
from wtforms.validators import InputRequired, Email, Length
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from flask_mail import Message
import xlsxwriter
import os
mail = Mail(app)

APP_ROOT = os.path.dirname(os.path.abspath(__file__))

def is_safe_url(target):
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and \
           ref_url.netloc == test_url.netloc

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

class LoginForm(FlaskForm):
    username = StringField('Имя пользователя', validators=[InputRequired(),Length(min=4,max=15)])
    password = PasswordField('Пароль',validators=[InputRequired(),Length(min=4, max=80)])

class RegisterForm(FlaskForm):
    email = StringField('email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=50)])
    username = StringField('username', validators=[InputRequired(), Length(min=4, max=15)])
    password = PasswordField('password', validators=[InputRequired(), Length(min=8, max=80)])


@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    form = LoginForm()
    if form.validate_on_submit():
        error = 'Неверное имя пользователя или пароль'
        user = Admin.query.filter_by(username=form.username.data).first()
        if user:
            print(user)
            if check_password_hash(user.password, form.password.data):
                login_user(user)
                return redirect(url_for('main'))
    return render_template('login.html', form=form, error=error)


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    form = RegisterForm()

    if form.validate_on_submit():
        hashed_password = generate_password_hash(form.password.data, method='sha256')
        new_user = Admin(username=form.username.data, email=form.email.data, password=hashed_password)
        db.session.add(new_user)
        db.session.commit()

        return redirect(url_for('main'))
        #return '<h1>' + form.username.data + ' ' + form.email.data + ' ' + form.password.data + '</h1>'

    return render_template('signup.html', form=form)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

if __name__ == "__main__":
    app.secret_key = os.urandom(12)
    app.run(debug=True,host='0.0.0.0', port=4000)


@app.route('/', methods=['GET'])
@login_required
def main():
    page =  request.args.get('page')
    if page and page.isdigit():
        page =  int(page)
    else:
        page = 1
    students = Student.query.order_by(Student.date_in.desc(), Student.surname)

    pages = students.paginate(page=page, per_page=5)
    return render_template('main.html', students=students.all(), pages=pages, name=current_user.username)


@app.route('/save', methods=['POST'])
@login_required
def save():
    target = os.path.join(APP_ROOT, 'static/')
    if request.files.getlist("image"):
        z = request.files.getlist("image")[0]
        z.filename = request.form['name'] +' ' + request.form['surname'] + request.form['id_people']
        randomm = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        filename = str(z.filename)
        destination = "/".join([target, filename])
        z.save(destination)
        f = '/static/' + filename
    else:
        destination = ''
        f = ''
    if 'id' in request.form:
        id = request.form['id']
        student = Student.query.get(id)
    else:
        student = Student(name=request.form['name'],
                          surname=request.form['surname'],
                          patronymic=request.form['patronymic'],
                          birth_date=request.form['birth_date'],
                          birth_place=request.form['birth_place'],
                          image=destination,
                          grp_id=request.form['grp_id'],
                          id_people=request.form['id_people'],
                          social_web_profile=request.form['social_web_profile'],
                          thesis_topic=request.form['thesis_topic'],
                          advisor=request.form['advisor'],
                          date_of_defense=request.form['date_of_defense'],
                          date_of_discharge=request.form['date_of_discharge'],
                          reason_of_discharge=request.form['reason_of_discharge'],
                          sex=request.form['sex'],
                          id_nationality=request.form['id_nationality'],
                          old_surname=request.form['old_surname'],
                          old_name=request.form['old_name'],
                          old_patronymic=request.form['old_patronymic'],
                          date_in=request.form['date_in'])
        print(student)

    student.name = request.form['name']
    student.surname = request.form['surname']
    student.patronymic = request.form['patronymic']
    student.birth_place = request.form['birth_place']
    if request.form['birth_date'] == '':
        student.birth_date = None
    else:
        student.birth_date = request.form['birth_date']
    student.grp_id = request.form['grp_id']
    student.social_web_profile = request.form['social_web_profile']
    student.thesis_topic = request.form['thesis_topic']
    student.advisor = request.form['advisor']
    if request.form['date_of_defense'] == '':
        student.date_of_defense = None
    else:
        student.date_of_defense = request.form['date_of_defense']
    if request.form['date_of_discharge'] == '':
        student.date_of_discharge = None
    else:
        student.date_of_discharge = request.form['date_of_discharge']
    student.reason_of_discharge = request.form['reason_of_discharge']
    student.sex = request.form['sex']
    student.id_nationality = request.form['id_nationality']
    student.old_surname = request.form['old_surname']
    student.old_name = request.form['old_name']
    student.old_patronymic = request.form['old_patronymic']
    if request.form['date_in'] == '':
        student.date_in = None
    else:
        student.date_in = request.form['date_in']
    student.image = f
    student.grp_id = request.form['grp_id']
    student.id_People = request.form['id_people']

    db.session.add(student)
    db.session.commit()

    return redirect(url_for('main'))


@app.route('/add')
@login_required
def add_student():
    groups = Grp.query.all()
    return render_template('student_form.html', groups=groups, name=current_user.username)


@app.route('/del_yes',methods=['POST'])
def delet():
    id = request.form['id']
    student = Student.query.get(id)
    db.session.delete(student)
    db.session.commit()
    return redirect(url_for('main'))


@app.route('/edit/<id>')
@login_required
def edit(id):
    student = Student.query.get(id)
    groups = Grp.query.all()
    return render_template('student_form.html', student=student, groups=groups, name=current_user.username)


@app.route('/search_results',methods=['POST'])
def show_results():
    search_word = request.form['word']
    students = Student.query.whoosh_search(search_word).all()
    groups = Grp.query.whoosh_search(search_word).all()
    return render_template('serchres.html', students=students, name=current_user.username, groups=groups)


@app.route('/stud_in_grp/<id>')
def ssig(id):
    grp = Grp.query.get(id)
    students = grp.students.all()
    return render_template('stofgrp.html', students=students, name=current_user.username, grp=grp)


@app.route('/registration',methods=['POST'])
def reg():
    return render_template('registration.html')


@app.route('/openfiledialog', methods=['POST','GET'])
def ofd():
    target = os.path.join(APP_ROOT, 'static/')
    if request.files.getlist("file"):
        z = request.files.getlist("file")[0]
        filename = z.filename
        destination = "/".join([target, filename])
        z.save(destination)
    else:
        filename = ''
    if filename == '':
        return redirect(url_for('main'))
    else:
        wb = load_workbook(filename = destination)
        sheet = wb['students']
        groups = {g.name: g for g in Grp.query.all()}
        count = 0
        for row in sheet.rows:
            if count == 0:
                count = 1
                continue
            groupname = row[1].value
            if groupname not in groups:
                g = Grp(id=row[0].value, id_plan=row[2].value, name=groupname)
                db.session.add(g)
                db.session.commit()
                groups[groupname] = g
            else:
                g = groups[groupname]
            s = row[8]
            if s.value == 'NULL':
                s.value=None
            student = Student(name= row[5].value,
                              surname= row[4].value,
                              patronymic= row[6].value,
                              birth_date= row[13].value,
                              birth_place= row[14].value,
                              grp_id=g.id,
                              id_people= row[3].value,
                              sex=row[12].value,
                              id_nationality= row[15].value,
                              old_surname= row[9].value ,
                              old_name= row[10].value,
                              old_patronymic= row[11].value,
                              date_in= row[7].value,
                              date_of_discharge = s.value,
                              advisor = '',
                              thesis_topic = '',
                              social_web_profile = '',
                              reason_of_discharge = row[16].value,
                              )
            db.session.add(student)
            db.session.commit()
        return redirect(url_for('main'))


@app.route("/openfiledialogpage", methods=['GET', 'POST'])
def ofdp():
    return render_template('openfiledialogpage.html', name=current_user.username)


@app.route('/report_generation')
def generation():
    return render_template('report_generation.html', name=current_user.username)


@app.route('/report_generate',methods=['POST','GET'])
def generate():
    workbook = xlsxwriter.Workbook('Отчет о выпускниках.xlsx')
    count = 0
    students = Student.query.all()
    worksheet = workbook.add_worksheet()
    if count == 0 and request.form.get('2001'):
        students = Student.query.filter_by(year_of_issue='2001')
        count += 1

    if count == 0 and request.form.get('2013'):
        students = Student.query.filter_by(year_of_issue='2013')
        count += 1
    if count > 0 and request.form.get('2013'):
        students = students.union(Student.query.filter_by(year_of_issue='2013'))

    if count == 0 and request.form.get('2002'):
        students = Student.query.filter_by(year_of_issue='2002')
        count += 1
    if count > 0 and request.form.get('2002'):
        students = students.union(Student.query.filter_by(year_of_issue='2002'))

    count = [1]*100
    letter = [65]*100
    Letter = 64
    for i in students:
        if request.form.get('1'):
            worksheet.write_column(chr(letter[0]) + str(count[0]),[i.surname, ])
            count[0] += 1
            if count[0] == 2:
                Letter += 1
                letter[0] = Letter
        if request.form.get('2'):
            worksheet.write_column(chr(letter[1]) + str(count[1]),[i.name, ])
            count[1] += 1
            if count[1] == 2:
                Letter += 1
                letter[1] = Letter
        if request.form.get('3'):
            worksheet.write_column(chr(letter[2]) + str(count[2]),[i.patronymic, ])
            count[2] += 1
            if count[2] == 2:
                Letter += 1
                letter[2] = Letter
        if request.form.get('4'):
            worksheet.write_column(chr(letter[3]) + str(count[3]),[i.year_of_issue, ])
            count[3] += 1
            if count[3] == 2:
                Letter += 1
                letter[3] = Letter
    workbook.close()
    return send_from_directory('C:\\Users\\Константин\\PycharmProjects\\blog','Отчет о выпускниках.xlsx')#redirect(url_for('main'))


@app.route('/change_password')
def change_pass():
    user = current_user
    return render_template('change_password.html', user=user, name=user.username, b=current_user.password)


@app.route('/new_pass', methods=['POST','GET'])
def new_pass():
    if check_password_hash(current_user.password, request.form['old_password']) :
        current_user.password = generate_password_hash(request.form['new_password'], method='sha256')
    db.session.commit()
    return render_template('change_password.html', user=current_user.username,a=generate_password_hash (request.form['old_password']))


@app.route('/forgot_password')
def forgot_pass():
    user = Admin.query.filter_by(username='spider-man').first()
    return render_template('forgot_password.html', user=user)


@app.route('/send_new_pass', methods=['POST','GET'])
def send_pass():
    user = Admin.query.filter_by(username=request.form['username']).first()
    password = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    user.password = generate_password_hash(password, method='sha256')
    db.session.commit()
    msg = Message('hello!', sender='olenov1994@gmail.com', recipients=['olenov1994@gmail.com'])
    msg.body = 'vash noviy password -' + password
    mail.send(msg)
    return render_template('forgot_password.html', user=user)


@app.route('/groups', methods=['GET'])
def groups():
    groups = Grp.query.all()
    return render_template('groups.html', groups=groups, name=current_user.username)


@app.route('/save_group', methods=['POST'])
def sg():
    if 'id' in request.form:
        id = request.form['id']
        group = Grp.query.get(id)
    else:
        group = Grp(id_plan=request.form['id_plan'],name=request.form['name'])

    group.name = request.form['name']
    group.id_plan = request.form['id_plan']

    db.session.add(group)
    db.session.commit()

    return redirect(url_for('groups'))


@app.route('/edit_group/<id>')
@login_required
def eg(id):
    student = Student.query.all()
    group = Grp.query.get(id)
    return render_template('group_form.html', student=student, group=group, name=current_user.username)


@app.route('/add_grp')
@login_required
def ag():
    return render_template('group_form.html', name=current_user.username)


@app.route('/del_group',methods=['POST'])
def delg():
    id = request.form['id']
    group = Grp.query.get(id)
    db.session.delete(group)
    db.session.commit()
    return redirect(url_for('groups'))


print('VIEWS:', app.root_path)